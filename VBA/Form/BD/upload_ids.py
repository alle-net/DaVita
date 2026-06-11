import openpyxl
import uuid
import os

# ============================================================
# VARIAVEIS - PREENCHA AQUI
# ============================================================
CAMINHO_PASTA = r"BD"               # pasta onde esta o arquivo
NOME_ARQUIVO = "Base Antiga.xlsx"   # nome do arquivo
NOME_ABA      = "Antiga"            # nome da aba/planilha
COLUNA_ID     = "A"                 # coluna do ID (A, B, C...)
# ============================================================

def gerar_uuids(caminho_arquivo, nome_aba, coluna_id):
    if not os.path.exists(caminho_arquivo):
        print(f"ERRO: Arquivo nao encontrado: {caminho_arquivo}")
        return False

    wb = openpyxl.load_workbook(caminho_arquivo)
    if nome_aba not in wb.sheetnames:
        print(f"ERRO: Aba '{nome_aba}' nao encontrada. Abas disponiveis: {wb.sheetnames}")
        return False

    ws = wb[nome_aba]
    col_idx = ord(coluna_id.upper()) - ord('A') + 1
    max_row = ws.max_row

    gerados = 0
    ja_existiam = 0

    for row in range(2, max_row + 1):
        celula = ws.cell(row=row, column=col_idx)
        valor = celula.value

        if valor is None or str(valor).strip() == '':
            celula.value = str(uuid.uuid4())
            gerados += 1
        else:
            ja_existiam += 1

    wb.save(caminho_arquivo)
    wb.close()

    print(f"Arquivo: {caminho_arquivo}")
    print(f"Aba: {nome_aba}")
    print(f"Coluna: {coluna_id}")
    print(f"Total de linhas: {max_row - 1}")
    print(f"UUIDs gerados: {gerados}")
    print(f"IDs ja existentes (preservados): {ja_existiam}")
    print("Status: OK")
    return True

if __name__ == "__main__":
    print("=== GERADOR DE ID (UUID v4) ===")
    print()
    caminho_completo = os.path.join(CAMINHO_PASTA, NOME_ARQUIVO)
    gerar_uuids(caminho_completo, NOME_ABA, COLUNA_ID)
