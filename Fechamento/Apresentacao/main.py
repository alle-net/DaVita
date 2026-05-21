import json
import csv
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
import openpyxl
import time
import os
import glob
from datetime import datetime

MESES = {
    "jan": "01", "fev": "02", "mar": "03", "abr": "04",
    "mai": "05", "jun": "06", "jul": "07", "ago": "08",
    "set": "09", "out": "10", "nov": "11", "dez": "12"
}

TIMEOUT_CARREGAMENTO = 600
TIMEOUT_DOWNLOAD = 1200

def aguardar_excel_visivel(driver, timeout=TIMEOUT_CARREGAMENTO):
    print(f"  Aguardando dados...", end=" ", flush=True)
    inicio = time.time()
    while time.time() - inicio < timeout:
        for btn in driver.find_elements(By.TAG_NAME, "button"):
            if "Excel" in btn.text and btn.is_displayed():
                print("pronto")
                return True
        time.sleep(2)
    print("TIMEOUT")
    return False

def clicar_excel(driver):
    for btn in driver.find_elements(By.TAG_NAME, "button"):
        if "Excel" in btn.text and btn.is_displayed():
            driver.execute_script("arguments[0].click();", btn)
            return True
    return False

def load_config():
    with open("config.json", "r", encoding="utf-8") as f:
        return json.load(f)

def converter_data(mes_abrev, ano):
    mes_abrev = mes_abrev.strip().lower()[:3]
    mes_num = MESES.get(mes_abrev)
    if not mes_num:
        raise ValueError(f"Mês desconhecido: '{mes_abrev}'")
    return f"{ano}-{mes_num}"

def preencher_campo_js(driver, nome_campo, valor):
    valor_json = json.dumps(valor)
    nome_json = json.dumps(nome_campo)
    driver.execute_script(f"""
        var el = document.querySelector('input[name={nome_json}]');
        if (!el) return;
        el.focus();
        var nativeInputValueSetter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
        nativeInputValueSetter.call(el, {valor_json});
        el.dispatchEvent(new Event('input', {{bubbles: true}}));
        el.dispatchEvent(new Event('change', {{bubbles: true}}));
        el.blur();
    """)

def preencher_data(driver, nome_campo, valor):
    preencher_campo_js(driver, nome_campo, valor)
    el = driver.find_element(By.NAME, nome_campo)
    if el.get_attribute("value") == valor:
        return True

    el.clear()
    el.click()
    el.send_keys(Keys.CONTROL + "a")
    el.send_keys(valor)
    el.send_keys(Keys.TAB)
    time.sleep(1)
    if el.get_attribute("value") == valor:
        return True

    actions = ActionChains(driver)
    actions.click(el).perform()
    el.send_keys(Keys.CONTROL + "a")
    el.send_keys(valor)
    el.send_keys(Keys.TAB)
    time.sleep(1)
    return el.get_attribute("value") == valor

def esperar_download(pasta_salvar, timeout=TIMEOUT_DOWNLOAD):
    inicio = time.time()
    while time.time() - inicio < timeout:
        temporarios = [f for f in os.listdir(pasta_salvar)
                       if f.endswith(('.crdownload', '.part', '.tmp'))
                       or f.startswith('~$')]
        if temporarios:
            time.sleep(1)
            continue
        completos = [f for f in os.listdir(pasta_salvar)
                     if not (f.endswith(('.crdownload', '.part', '.tmp'))
                             or f.startswith('~$'))]
        if completos:
            return True
        time.sleep(2)
    return False

def arquivos_existentes(pasta_salvar):
    existentes = set()
    if not os.path.exists(pasta_salvar):
        return existentes
    for padrao in ["*.xlsx", "*.xls"]:
        for f in glob.glob(os.path.join(pasta_salvar, padrao)):
            if not os.path.basename(f).startswith('~$'):
                existentes.add(os.path.normpath(os.path.abspath(f)))
    return existentes


def renomear_arquivo(pasta_salvar, nome_unidade, arquivos_antes, ja_renomeados):
    if not esperar_download(pasta_salvar):
        print(f"  AVISO: Timeout aguardando download")
        return False

    time.sleep(1)

    arquivos_depois = arquivos_existentes(pasta_salvar)
    if not arquivos_depois:
        print(f"  AVISO: Pasta vazia após download")
        return False

    novos = list(arquivos_depois - arquivos_antes)
    candidatos = [f for f in novos
                  if os.path.basename(f) not in ja_renomeados
                  and not os.path.basename(f).startswith('~$')]

    if not candidatos:
        candidatos = [f for f in arquivos_depois
                      if os.path.basename(f) not in ja_renomeados
                      and not os.path.basename(f).startswith('~$')]

    if not candidatos:
        print(f"  AVISO: Nenhum arquivo novo/encontrado para renomear")
        print(f"    Antes: {sorted(os.path.basename(f) for f in arquivos_antes)}")
        print(f"    Depois: {sorted(os.path.basename(f) for f in arquivos_depois)}")
        return False

    caminho_arquivo = sorted(candidatos, key=os.path.getmtime, reverse=True)[0]
    ext = os.path.splitext(caminho_arquivo)[1]
    novo_nome = f"{nome_unidade}{ext}"
    caminho_novo = os.path.join(pasta_salvar, novo_nome)

    if os.path.exists(caminho_novo):
        counter = 1
        while os.path.exists(f"{os.path.splitext(caminho_novo)[0]}_{counter}{ext}"):
            counter += 1
        caminho_novo = f"{os.path.splitext(caminho_novo)[0]}_{counter}{ext}"

    for tentativa in range(5):
        try:
            os.rename(caminho_arquivo, caminho_novo)
            break
        except PermissionError:
            print(f"  Arquivo ocupado, tentativa {tentativa+1}/5...")
            time.sleep(3)
    else:
        print(f"  ERRO: Não foi possível renomear após 5 tentativas")
        return False
    print(f"  Renomeado: {os.path.basename(caminho_arquivo)} -> {os.path.basename(caminho_novo)}")
    return os.path.basename(caminho_novo)

def gerar_resumo(pasta_salvar, arquivo_unidades, aba_unidades, coluna_unidade):
    print("\nGerando Resumo...")

    unidades = []
    mes_ref = ""
    ano_ref = ""
    wb_lista = openpyxl.load_workbook(arquivo_unidades, data_only=True)
    ws_lista = wb_lista[aba_unidades]
    header = [str(ws_lista.cell(row=1, column=c).value or "").strip() for c in range(1, ws_lista.max_column + 1)]
    col_unidade = header.index(coluna_unidade) + 1
    col_mes = next((i + 1 for i, h in enumerate(header) if h.lower().strip() == "mes"), None)
    col_ano = header.index("ano") + 1
    for r in range(2, ws_lista.max_row + 1):
        v = ws_lista.cell(row=r, column=col_unidade).value
        if v:
            unidades.append(str(v).strip())
        if not mes_ref and col_mes and col_ano:
            m = ws_lista.cell(row=r, column=col_mes).value
            a = ws_lista.cell(row=r, column=col_ano).value
            if m and a:
                mes_ref = str(m).strip().lower()[:3]
                if isinstance(a, (int, float)):
                    ano_ref = str(int(float(a)))
                else:
                    ano_ref = str(a).strip()
    wb_lista.close()

    if mes_ref:
        ref_formatada = converter_data(mes_ref, ano_ref)
        partes = ref_formatada.split("-")
        ref_texto = f"{partes[1]}/{partes[0]}"
    else:
        ref_texto = ""

    total_geral = 0.0
    unidade_valores = {}

    for padrao in ["*.xlsx", "*.xls"]:
        for arquivo in glob.glob(os.path.join(pasta_salvar, padrao)):
            base = os.path.basename(arquivo)
            if base.startswith("~$") or base == "Resumo.csv":
                continue
            nome_unidade = os.path.splitext(base)[0]
            try:
                wb = openpyxl.load_workbook(arquivo, data_only=True)
                ws = wb.active
                soma = 0.0
                for r in range(2, ws.max_row + 1):
                    val = ws.cell(row=r, column=6).value
                    if isinstance(val, (int, float)):
                        soma += val
                wb.close()
                unidade_valores[nome_unidade] = soma
                total_geral += soma
            except Exception as e:
                print(f"  AVISO: Erro ao ler {base}: {e}")

    caminho_csv = os.path.join(pasta_salvar, "Resumo.csv")
    with open(caminho_csv, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        now = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        writer.writerow(["RESUMO DO PROCESSO"])
        writer.writerow([now])
        writer.writerow([f"Referencia de: {ref_texto}"])
        writer.writerow(["Valor Total Faturado", f"R$ {total_geral:.2f}"])
        writer.writerow([])
        writer.writerow(["Unidade", "Valor Faturado (R$)"])
        for unidade in unidades:
            valor = unidade_valores.get(unidade, 0.0)
            writer.writerow([unidade, f"{valor:.2f}"])

    print(f"Resumo salvo em: {caminho_csv}")
    print(f"Valor total faturado: R$ {total_geral:.2f}")

def main():
    config = load_config()
    url = config["url"]
    usuario = config["usuario"]
    senha = config["senha"]
    arquivo_unidades = config["arquivo_unidades"]
    aba_unidades = config["aba_unidades"]
    coluna_unidade = config.get("coluna_unidade", "unidade")
    pasta_salvar = config["pasta_salvar"]
    pasta_downloads = os.path.join(os.environ["USERPROFILE"], "Downloads")
    pasta_salvar = os.path.join(pasta_downloads, f'{datetime.now().strftime("%Y%m%d")} Apresentacao')

    wb = openpyxl.load_workbook(arquivo_unidades, data_only=True)
    ws = wb[aba_unidades]

    header = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
    col_unidade = header.index(coluna_unidade) + 1
    col_mes = next((i + 1 for i, h in enumerate(header) if h.lower().strip() == "mes"), None)
    col_ano = header.index("ano") + 1
    if col_mes is None:
        raise ValueError("Coluna 'mes' não encontrada no cabeçalho")

    dados = []
    for r in range(2, ws.max_row + 1):
        unidade = ws.cell(row=r, column=col_unidade).value
        mes = ws.cell(row=r, column=col_mes).value
        ano = ws.cell(row=r, column=col_ano).value
        if not unidade or not mes or not ano:
            continue
        mes_str = str(mes).strip().lower()
        if isinstance(ano, (int, float)):
            ano_str = str(int(float(ano)))
        else:
            ano_str = str(ano).strip()
        data_formatada = converter_data(mes_str, ano_str)
        dados.append({"unidade": str(unidade).strip(), "mes_ano": data_formatada})

    print(f"Registros carregados: {len(dados)}")
    for d in dados:
        print(f"  {d['unidade']} | {d['mes_ano']}")

    os.makedirs(pasta_salvar, exist_ok=True)
    print(f"Pasta: '{pasta_salvar}'")

    options = webdriver.EdgeOptions()
    prefs = {
        "download.default_directory": os.path.abspath(pasta_salvar),
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
    }
    options.add_experimental_option("prefs", prefs)

    driver = webdriver.Edge(options=options)

    try:
        driver.get(url)
        wait = WebDriverWait(driver, 15)

        wait.until(EC.presence_of_element_located((By.NAME, "username"))).send_keys(usuario)
        driver.find_element(By.NAME, "password").send_keys(senha)
        driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()
        wait.until(EC.url_changes(url))
        print("Login efetuado!")
        time.sleep(3)

        wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Relat')]"))).click()
        time.sleep(2)
        driver.find_element(By.XPATH, "//*[contains(text(), 'Faturamento') and contains(text(), 'Mensal')]").click()
        time.sleep(5)

        base_url = driver.current_url
        print(f"URL base: {base_url}")

        ja_renomeados = set()

        for i, item in enumerate(dados):
            print(f"[{i + 1}/{len(dados)}] {item['unidade']} | {item['mes_ano']}")

            caminho_unidade = os.path.join(pasta_salvar, f"{item['unidade']}.xlsx")
            if os.path.exists(caminho_unidade):
                print(f"  JA EXISTE: Pulando...")
                continue

            try:
                driver.get(base_url)
                time.sleep(4)
                wait_local = WebDriverWait(driver, 20)
            except Exception:
                driver.refresh()
                time.sleep(4)
                wait_local = WebDriverWait(driver, 20)

            try:
                select_elem = wait_local.until(
                    EC.element_to_be_clickable((By.NAME, "empresa"))
                )
                Select(select_elem).select_by_visible_text(item["unidade"])
                print(f"  Unidade: '{item['unidade']}'")
            except Exception as e:
                print(f"  ERRO ao selecionar unidade: {e}")
                continue

            if not preencher_data(driver, "mes_de", item["mes_ano"]):
                print(f"  ERRO: campo mes_de não preenchido")
                continue
            if not preencher_data(driver, "mes_ate", item["mes_ano"]):
                print(f"  ERRO: campo mes_ate não preenchido")
                continue

            de_val = driver.find_element(By.NAME, "mes_de").get_attribute("value")
            ate_val = driver.find_element(By.NAME, "mes_ate").get_attribute("value")
            print(f"  mes_de='{de_val}' | mes_ate='{ate_val}'")

            for btn in driver.find_elements(By.TAG_NAME, "button"):
                if btn.text.strip() == "CONSULTAR":
                    driver.execute_script("arguments[0].click();", btn)
                    break

            if not aguardar_excel_visivel(driver):
                print(f"  ERRO: Timeout no carregamento dos dados")
                continue

            arquivos_antes = arquivos_existentes(pasta_salvar)
            if not clicar_excel(driver):
                print(f"  ERRO: Botão Excel não encontrado")
                continue
            print(f"  Download iniciado")

            nome_renomeado = renomear_arquivo(pasta_salvar, item["unidade"], arquivos_antes, ja_renomeados)
            if nome_renomeado:
                ja_renomeados.add(nome_renomeado)
                print(f"  OK: {item['unidade']}")
            else:
                print(f"  FALHA: {item['unidade']}")

        print("\n=== Processo concluído! ===")
        for f in os.listdir(pasta_salvar):
            print(f"  - {f}")

        gerar_resumo(pasta_salvar, arquivo_unidades, aba_unidades, coluna_unidade)

    except Exception as e:
        print(f"Erro: {e}")
        import traceback
        traceback.print_exc()
    finally:
        driver.quit()

if __name__ == "__main__":
    main()
